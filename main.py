import time

from openpyxl import Workbook, load_workbook, utils
from tabulate import tabulate

from helpers import clean_list_tuples, get_tuples_between_tags, listing_sheet, transform_tuples

tcpu0 = time.time()

wb = load_workbook(filename='original_vae.xlsx', read_only=True, data_only=True)

# Inicialización de data
listing_data = []  # Aquí se almacena todos los apus
equipment_list = labour_list = materials_list = []

# Bucle busqueda y limpieza de insumos
for sheet in wb.sheetnames:
    if sheet == 'Rubros':
        continue

    _detailedSheet = listing_sheet(wb[sheet], 1, 11)

    equipment_list = equipment_list + clean_list_tuples(get_tuples_between_tags(
        _detailedSheet, 'EQUIPOS', 'MANO DE OBRA', True), (0, 7, 8))
    labour_list = labour_list + clean_list_tuples(get_tuples_between_tags(
        _detailedSheet, 'MANO DE OBRA', 'MATERIALES', True), (0, 7, 8))
    materials_list = materials_list + clean_list_tuples(get_tuples_between_tags(
        _detailedSheet, 'MATERIALES', 'TRANSPORTE', True), (0, 2, 7, 8))
    
    # print(tabulate(equipment_list))

# Limpieza de repetidos y ordenamiento
clean_equipment_dict = {index: value for index,
                        value in enumerate(sorted(set(equipment_list)), 77)}
clean_labour_dict = {index: value for index,
                     value in enumerate(sorted(set(labour_list)), 532)}
clean_materials_dict = {index: value for index,
                        value in enumerate(sorted(set(materials_list)), 222)}

print(tabulate(clean_equipment_dict.items()))
print(tabulate(clean_labour_dict.items()))
print(tabulate(clean_materials_dict.items()))

# # Segunda vuelta de rubros
# for sheet in wb.sheetnames:
#     if sheet == 'Rubros':
#         continue

#     _detailedSheet = listing_sheet(wb[sheet],1,6)
#     # print(_detailedSheet)

#     _equipment_list = clean_list_tuples(get_tuples_between_tags(
#         _detailedSheet, 'EQUIPOS', 'MANO DE OBRA', True), (0, 1, 2))
#     _new_equipment_list = transform_tuples(
#         clean_equipment_dict, _equipment_list)

#     _labour_list = clean_list_tuples(get_tuples_between_tags(
#         _detailedSheet, 'MANO DE OBRA', 'MATERIALES', True), (0, 1, 2))
#     _new_labour_list = transform_tuples(clean_labour_dict, _labour_list)

#     _materials_list = clean_list_tuples(get_tuples_between_tags(
#         _detailedSheet, 'MATERIALES', 'TRANSPORTE', True), (0, 2, 3, 4))
#     _new_materials_list = transform_tuples(
#         clean_materials_dict, _materials_list)

#     _new_pa = {'SHEET': sheet, 'ITEM': _detailedSheet[8][0], 'RUBRO': (_detailedSheet[4][0]).replace('RUBRO: ',''), 'UNIDAD': _detailedSheet[6][0].replace('UNIDAD: ',''),   'EQUIPO': _new_equipment_list,
#                'MANO DE OBRA': _new_labour_list, 'MATERIALES': _new_materials_list}
#     listing_data.append(_new_pa)
    

# # Creación del diccionario general de Rubros
# dict_data = {index: value for index,
#              value in enumerate(listing_data, 164)}
# print(dict_data)

# Close the workbook after reading
wb.close()


# Crear un nuevo libro de Excel y EXPORTA los resultados
excel_book = Workbook()
excel_book.remove(excel_book['Sheet'])
_active_sheet = excel_book.create_sheet(title='EQUIPOS')
for index, value in enumerate(clean_equipment_dict.items(), start=1):
    try:
        _active_sheet[f'A{index}'] = value[0]
        _active_sheet[f'B{index}'] = value[1][0]
        _active_sheet[f'C{index}'] = value[1][1]
        _active_sheet[f'D{index}'] = value[1][2]
        # print(index, value)
    except IndexError:
        continue

_active_sheet = excel_book.create_sheet(title='MANO DE OBRA')
for index, value in enumerate(clean_labour_dict.items(), start=1):
    try:     
        _active_sheet[f'A{index}'] = value[0]
        _active_sheet[f'B{index}'] = value[1][0]
        _active_sheet[f'C{index}'] = value[1][1]
        _active_sheet[f'D{index}'] = value[1][2]
    except IndexError:
        continue

_active_sheet = excel_book.create_sheet(title='MATERIALES')
for index, value in enumerate(clean_materials_dict.items(), start=1):
    try:
        _active_sheet[f'A{index}'] = value[0]
        _active_sheet[f'B{index}'] = value[1][0]
        _active_sheet[f'C{index}'] = value[1][1]
        _active_sheet[f'D{index}'] = value[1][2]
        _active_sheet[f'E{index}'] = value[1][3]
    except IndexError:
        continue

# _active_sheet = excel_book.create_sheet(title='RUBROS')
# MAX_ITERATIONS = 20

# for index, (key, data) in enumerate(dict_data.items(), start=1):
#     _active_sheet[f'A{index}'] = key
#     _active_sheet[f'E{index}'] = data['RUBRO']
#     _active_sheet[f'F{index}'] = data['ITEM']
#     _active_sheet[f'G{index}'] = data['UNIDAD']
#     _active_sheet[f'H{index}'] = 1

#     for category, col_start in zip(['MATERIALES', 'MANO DE OBRA', 'EQUIPO'], [9, 49, 89]):
#         for _i, item in enumerate(data[category], start=1):
#             col_step = 2
#             if _i > MAX_ITERATIONS:
#                 break
#             try:
#                 col_letter_1 = utils.get_column_letter(
#                     col_start + (_i - 1) * col_step)
#                 col_letter_2 = utils.get_column_letter(
#                     col_start + (_i - 1) * col_step + 1)
#                 _active_sheet[f'{col_letter_1}{index}'] = float(item[0])
#                 try:
#                     _active_sheet[f'{col_letter_2}{index}'] = float(item[1])
#                 except ValueError:
#                     continue
#             except IndexError:
#                 continue

# Guardar el libro de Excel
excel_book.save("output_vae.xlsx")

print("Se ha creado el archivo Excel: output.xlsx")

print('Finalizado en: ', (time.time()-tcpu0), 'segundos')
